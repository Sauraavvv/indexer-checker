"""
processing.py — Core URL index-checking logic (multi-sheet, parallel workers).
Imported by app.py (Streamlit UI) or run standalone via run_sheets().
"""

import asyncio
import csv
import io
import logging
import os
import re
import time as _time
from datetime import datetime
from typing import Callable, Optional

import gspread
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import async_playwright


# ─────────────────────────────────────────────────────────────
# CONFIG  (edit these if paths change)
# ─────────────────────────────────────────────────────────────
URLS_PER_TAB          = 25    # rows per batch tab — keep ≤ 30 (one viewport)
MAX_BATCHES_PER_SHEET = 0   # 0 = no limit, batches auto = ceil(total_urls / 25)
DEFAULT_MASTER_TAB    = 1     # 1-based tab index containing the URL list

CREDENTIALS_FILE = os.path.join(
    os.path.dirname(__file__),
    "client_secret_578289306233-ed187otkn67u5i8iblk70f20krcr9vgs.apps.googleusercontent.com.json",
)
TOKEN_FILE  = os.path.join(os.path.dirname(__file__), "token.json")
SCOPES      = ["https://www.googleapis.com/auth/spreadsheets"]
PROFILE_DIR = os.path.join(os.path.dirname(__file__), "browser_profile")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "results.xlsx")

COL_HEADER_H = 21
ROW_H        = 21
ROW_NUM_W    = 46
COL_A_W      = 100
# ─────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────
_LOG_DIR = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(_LOG_DIR, exist_ok=True)

def _make_log_file() -> str:
    return os.path.join(_LOG_DIR, f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

_log_file = _make_log_file()
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s.%(msecs)03d  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.FileHandler(_log_file, encoding="utf-8")],
)
_log = logging.getLogger("IndexChecker")

def _dbg(tag: str, msg: str):
    _log.debug("[%s] %s", tag, msg)


# ─────────────────────────────────────────────────────────────
# URL HELPERS
# ─────────────────────────────────────────────────────────────
def normalize_url(u: str) -> str:
    u = u.lower().strip()
    u = re.sub(r'^https?://(www\.)?', '', u)
    return u.rstrip('/')

def extract_sheet_id(url: str) -> str:
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        raise ValueError(f"Sheet ID not found in URL: {url}")
    return m.group(1)


# ─────────────────────────────────────────────────────────────
# GSPREAD AUTH + HELPERS
# ─────────────────────────────────────────────────────────────
def _api_call(fn, *args, **kwargs):
    """Retry gspread calls with exponential backoff on 429."""
    delay = 5
    for attempt in range(6):
        try:
            return fn(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            if "429" in str(e) and attempt < 5:
                print(f"   ⏳ Rate limited — waiting {delay}s...")
                _time.sleep(delay)
                delay = min(delay * 2, 60)
            else:
                raise


def get_gspread_client() -> gspread.Client:
    """Authorize gspread via OAuth2, caching token.json."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                raise FileNotFoundError(
                    f"credentials.json not found: {CREDENTIALS_FILE}\n"
                    "Download from Google Cloud Console → APIs & Services → Credentials."
                )
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as fh:
            fh.write(creds.to_json())

    gc = gspread.authorize(creds)
    try:
        import urllib.request
        import json as _j
        req = urllib.request.Request(
            "https://www.googleapis.com/oauth2/v1/userinfo",
            headers={"Authorization": f"Bearer {creds.token}"},
        )
        info = _j.loads(urllib.request.urlopen(req).read())
        print(f"   gspread authorized as: {info.get('email', 'unknown')}")
    except Exception:
        pass
    return gc


# ─────────────────────────────────────────────────────────────
# GOOGLE SHEETS API — TAB OPERATIONS
# ─────────────────────────────────────────────────────────────
def create_batch_tab_api(gc: gspread.Client, sheet_id: str,
                         tab_name: str, urls: list[str]) -> str:
    """Create a worksheet named tab_name, write URLs, return gid."""
    _dbg("TAB", f"[API] Creating '{tab_name}' ({len(urls)} URLs)...")
    sh = _api_call(gc.open_by_key, sheet_id)
    try:
        old = _api_call(sh.worksheet, tab_name)
        _api_call(sh.del_worksheet, old)
    except gspread.exceptions.WorksheetNotFound:
        pass
    ws  = _api_call(sh.add_worksheet, title=tab_name, rows=max(len(urls) + 5, 30), cols=3)
    gid = str(ws.id)
    _api_call(ws.update, [["URL", "Status", "Checked At"]] + [[u] for u in urls],
              "A1", value_input_option="RAW")
    _dbg("TAB", f"[API] '{tab_name}' ready — gid={gid}")
    return gid


def write_batch_results_sync(gc: gspread.Client, sheet_id: str,
                             tab_name: str, results: list) -> None:
    """Write status + checked_at to col B/C of a batch worksheet."""
    try:
        sh   = _api_call(gc.open_by_key, sheet_id)
        ws   = _api_call(sh.worksheet, tab_name)
        data = [[r["status"], r["checked_at"]] for r in results]
        if data:
            _api_call(ws.update, data, f"B2:C{1 + len(data)}", value_input_option="RAW")
        _dbg("SHEET", f"Results written to '{tab_name}' ({len(data)} rows)")
    except Exception as e:
        print(f"   ⚠️  write_batch_results failed for '{tab_name}': {e}")


def write_master_results_sync(gc: gspread.Client, sheet_id: str,
                              master_gid: str, results: list,
                              start_row: int) -> None:
    """Write status + checked_at to col B/C of the master tab."""
    try:
        sh        = _api_call(gc.open_by_key, sheet_id)
        master_ws = next(
            (ws for ws in _api_call(sh.worksheets) if str(ws.id) == str(master_gid)), None
        )
        if not master_ws:
            return
        sorted_r  = sorted(results, key=lambda r: r["row_num"])
        data      = [[r["status"], r["checked_at"]] for r in sorted_r]
        row_start = start_row + 1
        if data:
            _api_call(master_ws.update, data,
                      f"B{row_start}:C{row_start + len(data) - 1}",
                      value_input_option="RAW")
        print(f"   ✅ Master tab updated ({len(data)} rows)")
    except Exception as e:
        print(f"   ⚠️  write_master_results failed: {e}")


# ─────────────────────────────────────────────────────────────
# EXCEL SAVE  (one workbook, one sheet per input sheet + All Results)
# ─────────────────────────────────────────────────────────────
def save_excel(results: list, output_file: str = OUTPUT_FILE) -> None:
    from collections import defaultdict
    wb       = Workbook()
    default  = wb.active
    default.title = "All Results"
    _fill_ws(default, sorted(results, key=lambda r: r["row_num"]))

    by_sheet: dict = defaultdict(list)
    for r in results:
        by_sheet[r.get("sheet_label", "Sheet_1")].append(r)

    for label, rows in sorted(by_sheet.items()):
        ws = wb.create_sheet(label[:31])
        _fill_ws(ws, sorted(rows, key=lambda r: r["row_num"]))

    wb.save(output_file)


def _fill_ws(ws, results: list) -> None:
    headers  = ["#", "URL", "Status", "Sheet", "Batch", "Checked At"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font

    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")

    for i, r in enumerate(results, 2):
        fill = green if "✅" in r["status"] else (red if "❌" in r["status"] else yellow)
        ws.cell(row=i, column=1, value=r["row_num"]).fill        = fill
        ws.cell(row=i, column=2, value=r["url"]).fill            = fill
        ws.cell(row=i, column=3, value=r["status"]).fill         = fill
        ws.cell(row=i, column=4, value=r.get("sheet_label","")).fill = fill
        ws.cell(row=i, column=5, value=r["batch"]).fill          = fill
        ws.cell(row=i, column=6, value=r["checked_at"]).fill     = fill

    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["F"].width = 22


# ─────────────────────────────────────────────────────────────
# TOOLTIP DISMISS
# ─────────────────────────────────────────────────────────────
async def wait_for_no_tooltip(page, timeout_ms: int = 5000, row_num: int = 0):
    await page.mouse.move(700, 30)
    await page.wait_for_timeout(150)
    await page.keyboard.press("Escape")
    await page.wait_for_timeout(200)
    for frame in page.frames:
        try:
            tip = frame.locator('.waffle-multilink-tooltip')
            if await tip.count() > 0:
                try:
                    await tip.wait_for(state="hidden", timeout=timeout_ms)
                except Exception:
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(600)
                    try:
                        await tip.wait_for(state="hidden", timeout=2000)
                    except Exception as e2:
                        _dbg("DISMISS", f"Could not dismiss: {e2}")
                break
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────
# TOOLTIP DETECT
# ─────────────────────────────────────────────────────────────
async def detect_tooltip(page, expected_url: str = "", row_num: int = 0,
                         quiet: bool = False) -> str:
    def _d(tag, msg):
        if not quiet:
            _dbg(tag, msg)

    for i, frame in enumerate(page.frames):
        try:
            tooltip = frame.locator('.waffle-multilink-tooltip')
            if await tooltip.count() == 0:
                continue

            _d("DETECT", f"frame[{i}]: tooltip present")

            if expected_url:
                for sel in ['.docs-linkbubble-url', '[class*="linkbubble-url"]',
                             '[class*="link-preview-url"]']:
                    el = frame.locator(sel)
                    if await el.count() > 0:
                        try:
                            raw = (await el.first.inner_text()).strip()
                            if raw and normalize_url(raw) != normalize_url(expected_url):
                                return "⚠️ Stale tooltip"
                        except Exception:
                            pass
                        break

            if expected_url:
                try:
                    _raw = (await tooltip.first.inner_text()).strip()
                    for _line in _raw.split("\n"):
                        _line = _line.strip()
                        if _line.startswith("http://") or _line.startswith("https://"):
                            if normalize_url(_line) != normalize_url(expected_url):
                                return "⚠️ Stale tooltip"
                            break
                except Exception:
                    pass

            for sel in ['.docs-linkbubble-preview-title', '[class*="preview-title"]',
                        '[class*="linkbubble-preview"] [class*="title"]',
                        '.appsElementsLinkPreview img',
                        '[class*="linkbubble"] img[src^="http"]']:
                el = frame.locator(sel)
                if await el.count() > 0:
                    return "✅ Indexed"

            for sel in ['svg[aria-label*="globe" i]', '[class*="default-favicon"]',
                        '[class*="globe"]', '[class*="generic-favicon"]',
                        '.docs-linkbubble-url-favicon']:
                el = frame.locator(sel)
                if await el.count() > 0 and await el.first.is_visible():
                    return "❌ Not Indexed"

            try:
                text = (await tooltip.first.inner_text()).strip()
            except Exception:
                return "⚠️ Popup (unclassified)"

            first_line = text.split("\n")[0].strip()
            if expected_url and (first_line.startswith("http://") or
                                  first_line.startswith("https://")):
                if normalize_url(first_line) != normalize_url(expected_url):
                    return "⚠️ Stale tooltip"
                return "❌ Not Indexed"
            if first_line.startswith("http://") or first_line.startswith("https://"):
                return "❌ Not Indexed"
            return f"⚠️ Popup: {text[:80]}"

        except Exception as e:
            _d("DETECT", f"frame[{i}] error: {e}")
            continue

    return "⚠️ No popup detected"


# ─────────────────────────────────────────────────────────────
# NAVIGATE TO TAB
# ─────────────────────────────────────────────────────────────
async def go_to_tab(page, sheet_id: str, gid: str):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={gid}"
    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
    await page.wait_for_timeout(3000)
    canvas_box = await page.evaluate("""() => {
        const c = [...document.querySelectorAll('canvas')]
            .sort((a,b)=>(b.width*b.height)-(a.width*a.height))[0];
        if (!c) return null;
        const r = c.getBoundingClientRect();
        return {x:r.x, y:r.y, w:r.width, h:r.height};
    }""")
    if canvas_box:
        await page.mouse.click(canvas_box['x'] + canvas_box['w'] / 2,
                                canvas_box['y'] + canvas_box['h'] / 2)
    await page.wait_for_timeout(300)


# ─────────────────────────────────────────────────────────────
# CALIBRATE ROW 1 Y
# ─────────────────────────────────────────────────────────────
async def calibrate_row1_y(page, cx, cy, cw, ch, cell_x) -> float:
    print("⏳ Calibrating row 1 pixel position...")
    await page.keyboard.press("Meta+ArrowUp")
    await page.wait_for_timeout(200)
    await page.keyboard.press("Meta+ArrowLeft")
    await page.wait_for_timeout(400)

    row1_y = None
    for offset in range(25, 120, 4):
        test_y = cy + offset
        await page.mouse.move(cell_x, test_y)
        await page.wait_for_timeout(700)
        for frame in page.frames:
            try:
                tip = frame.locator('.waffle-multilink-tooltip')
                if await tip.count() > 0 and await tip.first.is_visible():
                    row1_y = test_y
                    break
            except Exception:
                pass
        if row1_y:
            break
        await page.mouse.move(cx + 5, cy + 3)
        await page.wait_for_timeout(300)

    await wait_for_no_tooltip(page, row_num=0)

    if row1_y:
        print(f"   ✅ Row 1 Y = {row1_y:.0f}px")
    else:
        row1_y = cy + COL_HEADER_H + ROW_H / 2
        print(f"   ⚠️  Calibration fallback Y = {row1_y:.0f}px")

    await page.keyboard.press("Meta+ArrowUp")
    await page.wait_for_timeout(200)
    await page.keyboard.press("Meta+ArrowLeft")
    await page.wait_for_timeout(400)
    return row1_y


# ─────────────────────────────────────────────────────────────
# FETCH URLS FROM MASTER TAB
# ─────────────────────────────────────────────────────────────
async def fetch_master_urls(context, sheet_url: str,
                            tab_number: int) -> tuple[list[str], int]:
    """Returns (urls, start_row) — start_row=0 no header, 1=has header."""
    sheet_id = extract_sheet_id(sheet_url)
    page     = context.pages[0]

    gid = await page.evaluate("""(tabNum) => {
        const tabs = [...document.querySelectorAll('[id^="sheet-button-"]')];
        if (!tabs.length) return null;
        const el = tabs[tabNum - 1];
        if (!el) return null;
        const m = el.id.match(/sheet-button-(\\d+)/);
        return m ? m[1] : null;
    }""", tab_number)

    csv_url = (f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
               if gid else
               f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv")

    for attempt in range(1, 4):
        resp = await context.request.get(csv_url)
        if resp.ok:
            break
        await asyncio.sleep(3)
    else:
        raise RuntimeError(f"CSV fetch failed for {sheet_url}")

    content   = await resp.text()
    rows      = list(csv.reader(io.StringIO(content)))
    first_val = rows[0][0].strip() if rows else ""
    start_row = 0 if first_val.lower().startswith("http") else 1
    urls      = [r[0].strip() for r in rows[start_row:] if r and r[0].strip()]
    return urls, start_row


# ─────────────────────────────────────────────────────────────
# PROCESS ONE BATCH TAB
# ─────────────────────────────────────────────────────────────
async def process_batch(page, batch_urls: list[str], batch_num: int,
                        global_row_offset: int, row1_center_y: float,
                        cell_x: float, worker_id: int = 0,
                        sheet_label: str = "", sheet_idx: int = 0,
                        tab_name: str = "") -> list[dict]:
    results   = []
    row_count = len(batch_urls)
    W         = f"W{worker_id}"
    _label    = tab_name or f"Batch_{batch_num}"

    print(f"\n[{W}] {sheet_label} | {_label} — {row_count} URLs")

    await page.keyboard.press("Meta+ArrowUp")
    await page.wait_for_timeout(200)
    await page.keyboard.press("Meta+ArrowLeft")
    await page.wait_for_timeout(500)

    for row_idx in range(row_count):
        row_num    = row_idx + 1
        global_num = global_row_offset + row_idx
        url_text   = batch_urls[row_idx]

        print(f"[{W}] {_label} R{row_num}/{row_count} (#{global_num}): {url_text[:65]}")

        if row_idx > 0:
            await page.keyboard.press("ArrowDown")
            await page.wait_for_timeout(400)

        cell_y = row1_center_y + (ROW_H / 2) + row_idx * ROW_H

        await wait_for_no_tooltip(page, row_num=row_num)
        await page.wait_for_timeout(300)

        any_left = False
        for frame in page.frames:
            try:
                tip = frame.locator('.waffle-multilink-tooltip')
                if await tip.count() > 0 and await tip.first.is_visible():
                    any_left = True
                    break
            except Exception:
                pass
        if any_left:
            await wait_for_no_tooltip(page, row_num=row_num)
            await page.wait_for_timeout(500)

        MAX_ATTEMPTS  = 3
        MAX_WAIT_MS   = 2800
        POLL_EVERY_MS = 200
        INITIAL_MS    = 400
        status            = "⚠️ No popup detected"
        cell_y_correction = 0

        for attempt in range(1, MAX_ATTEMPTS + 1):
            hover_y = cell_y + cell_y_correction
            await page.mouse.move(700, hover_y)
            await page.mouse.move(cell_x, hover_y)

            await page.wait_for_timeout(INITIAL_MS)
            elapsed = INITIAL_MS
            status  = "⚠️ No popup detected"

            while elapsed < MAX_WAIT_MS:
                status = await detect_tooltip(page, expected_url=url_text,
                                              row_num=row_num, quiet=True)
                if status in ("✅ Indexed", "⚠️ Stale tooltip"):
                    break
                await page.wait_for_timeout(POLL_EVERY_MS)
                elapsed += POLL_EVERY_MS

            if status != "⚠️ Stale tooltip":
                break
            await wait_for_no_tooltip(page, row_num=row_num)
            await page.wait_for_timeout(200)
            cell_y_correction += ROW_H

        checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{W}]   → {status}")

        results.append({
            "row_num":     global_num,
            "url":         url_text,
            "status":      status,
            "batch":       _label,
            "sheet_label": sheet_label,
            "sheet_idx":   sheet_idx,
            "checked_at":  checked_at,
        })

        await wait_for_no_tooltip(page, row_num=row_num)
        await page.wait_for_timeout(400)

    return results


# ─────────────────────────────────────────────────────────────
# SINGLE-BATCH WORKER
# ─────────────────────────────────────────────────────────────
async def run_one_batch(worker_id: int, page,
                        tab_name: str, batch_urls: list[str],
                        sheet_id: str, gid: str, global_offset: int,
                        batch_num: int, gc: gspread.Client,
                        row1_center_y: float, cell_x: float,
                        all_results: list, lock: asyncio.Lock,
                        sheet_label: str = "", sheet_idx: int = 0) -> None:
    await go_to_tab(page, sheet_id, gid)

    canvas_box = await page.evaluate("""() => {
        const c = [...document.querySelectorAll('canvas')]
            .sort((a,b)=>(b.width*b.height)-(a.width*a.height))[0];
        if (!c) return null;
        const r = c.getBoundingClientRect();
        return {x:r.x, y:r.y, w:r.width, h:r.height};
    }""")
    if canvas_box:
        await page.mouse.click(canvas_box['x'] + canvas_box['w'] / 2,
                                canvas_box['y'] + canvas_box['h'] / 2)
    await page.wait_for_timeout(400)

    results = await process_batch(
        page, batch_urls, batch_num, global_offset,
        row1_center_y, cell_x,
        worker_id=worker_id, sheet_label=sheet_label,
        sheet_idx=sheet_idx, tab_name=tab_name,
    )

    async with lock:
        await asyncio.to_thread(write_batch_results_sync, gc, sheet_id, tab_name, results)
        all_results.extend(results)
        save_excel(all_results)
        print(f"\n[W{worker_id}] ✅ {tab_name} done — total so far: {len(all_results)}")


# ─────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────
async def run_sheets(
    sheet_urls: list[str],
    workers: int,
    master_tab: int = DEFAULT_MASTER_TAB,
    progress_cb: Optional[Callable[[int, int], None]] = None,
) -> list[dict]:
    """
    Process up to 5 sheets simultaneously with a shared worker pool.

    Args:
        sheet_urls  : list of Google Sheets URLs (1–5)
        workers     : number of parallel browser tabs
        master_tab  : 1-based tab number containing the URL list
        progress_cb : optional callback(batches_done, total_batches)

    Returns:
        Combined list of result dicts with keys:
        row_num, url, status, batch, sheet_label, sheet_idx, checked_at
    """
    valid_urls = [u.strip() for u in sheet_urls if u.strip()]
    if not valid_urls:
        return []

    gc          = get_gspread_client()
    all_results: list = []

    async with async_playwright() as pw:
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            channel="chrome",
            headless=False,
            viewport={"width": 1400, "height": 900},
        )

        # ── Open first sheet for canvas geometry + URL fetching ──────────
        page0 = await context.new_page()
        await page0.goto(valid_urls[0], wait_until="domcontentloaded", timeout=30000)
        await page0.wait_for_timeout(4000)

        # ── Spin up remaining worker pages ───────────────────────────────
        pages = [page0]
        for _ in range(1, workers):
            p = await context.new_page()
            await p.goto("about:blank")
            pages.append(p)

        # ── Canvas geometry (same across all sheets) ─────────────────────
        canvas_box = await page0.evaluate("""() => {
            const c = [...document.querySelectorAll('canvas')]
                .sort((a,b)=>(b.width*b.height)-(a.width*a.height))[0];
            if (!c) return null;
            const r = c.getBoundingClientRect();
            return {x:r.x, y:r.y, w:r.width, h:r.height};
        }""")
        if not canvas_box:
            await context.close()
            raise RuntimeError("Canvas not found — is the sheet loaded?")

        cx, cy = canvas_box['x'], canvas_box['y']
        cw, ch = canvas_box['w'], canvas_box['h']
        cell_x = cx + ROW_NUM_W + COL_A_W / 2

        # ── Collect batch items + sheet metadata from all sheets ─────────
        all_batch_items: list[dict] = []
        sheet_meta:      dict       = {}

        for sheet_idx, sheet_url in enumerate(valid_urls):
            sheet_id    = extract_sheet_id(sheet_url)
            sheet_label = f"Sheet_{sheet_idx + 1}"
            print(f"\n📄 Loading {sheet_label}: {sheet_url[:70]}...")

            await page0.goto(sheet_url, wait_until="domcontentloaded", timeout=30000)
            await page0.wait_for_timeout(4000)

            master_gid = await page0.evaluate("""(tabNum) => {
                const tabs = [...document.querySelectorAll('[id^="sheet-button-"]')];
                if (!tabs.length) return null;
                const el = tabs[tabNum - 1];
                if (!el) return null;
                const m = el.id.match(/sheet-button-(\\d+)/);
                return m ? m[1] : null;
            }""", master_tab)

            if master_gid:
                await page0.goto(
                    f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit#gid={master_gid}",
                    wait_until="domcontentloaded", timeout=30000,
                )
                await page0.wait_for_timeout(3000)

            urls, start_row = await fetch_master_urls(context, sheet_url, master_tab)
            # 0 = no limit → process all URLs, batches auto-calculated
            if MAX_BATCHES_PER_SHEET > 0:
                urls = urls[:MAX_BATCHES_PER_SHEET * URLS_PER_TAB]
            batches         = [urls[i:i + URLS_PER_TAB]
                               for i in range(0, len(urls), URLS_PER_TAB)]

            print(f"   ✅ {len(urls)} URLs → {len(batches)} batches")
            sheet_meta[sheet_idx] = {
                "sheet_id":  sheet_id,
                "master_gid": master_gid,
                "start_row": start_row,
                "label":     sheet_label,
            }

            for batch_num, batch in enumerate(batches, 1):
                tab_name = f"S{sheet_idx + 1}_Batch_{batch_num}"
                all_batch_items.append({
                    "sheet_idx":    sheet_idx,
                    "sheet_id":     sheet_id,
                    "sheet_label":  sheet_label,
                    "batch_num":    batch_num,
                    "batch_urls":   batch,
                    "tab_name":     tab_name,
                    "gid":          None,
                    "global_offset": (batch_num - 1) * URLS_PER_TAB + 1,
                })

        total_batches = len(all_batch_items)
        total_rounds  = (total_batches + workers - 1) // workers
        print(f"\n📊 {len(valid_urls)} sheets | {total_batches} total batches | "
              f"{workers} workers | {total_rounds} rounds")

        # ── Round-based: create + process `workers` batches per round ────
        lock          = asyncio.Lock()
        row1_center_y = None
        done_count    = 0

        for round_num in range(total_rounds):
            round_items = all_batch_items[round_num * workers:(round_num + 1) * workers]
            n           = len(round_items)

            print(f"\n{'='*60}")
            print(f"📦 Round {round_num+1}/{total_rounds} — creating {n} tab(s)...")

            for item in round_items:
                item["gid"] = await asyncio.to_thread(
                    create_batch_tab_api, gc,
                    item["sheet_id"], item["tab_name"], item["batch_urls"],
                )

            # Calibrate once on the very first batch tab created
            if row1_center_y is None and round_items:
                first = round_items[0]
                await go_to_tab(page0, first["sheet_id"], first["gid"])
                await page0.mouse.click(cx + cw / 2, cy + ch / 2)
                await page0.wait_for_timeout(400)
                row1_center_y = await calibrate_row1_y(page0, cx, cy, cw, ch, cell_x)

            print(f"   🔍 Processing {n} batch(es) in parallel...")
            tasks = [
                run_one_batch(
                    j, pages[j],
                    item["tab_name"], item["batch_urls"],
                    item["sheet_id"], item["gid"], item["global_offset"],
                    item["batch_num"], gc, row1_center_y, cell_x,
                    all_results, lock,
                    sheet_label=item["sheet_label"], sheet_idx=item["sheet_idx"],
                )
                for j, item in enumerate(round_items)
            ]
            await asyncio.gather(*tasks)

            done_count += n
            if progress_cb:
                progress_cb(done_count, total_batches)
            print(f"   ✅ Round {round_num+1} complete — {done_count}/{total_batches} batches done")

        # ── Write results back to each master tab ────────────────────────
        print("\n📝 Writing results to master tabs...")
        for sheet_idx, meta in sheet_meta.items():
            sheet_results = [r for r in all_results if r["sheet_idx"] == sheet_idx]
            if sheet_results:
                await asyncio.to_thread(
                    write_master_results_sync, gc,
                    meta["sheet_id"], meta["master_gid"],
                    sheet_results, meta["start_row"],
                )

        await context.close()

    save_excel(all_results)
    return all_results
